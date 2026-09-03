#!/usr/bin/env python3
"""Filter a bounded Google Sheets scan or explicit XLSX to the regular-work window.

The connector must read FORMATTED_VALUE cells.  This script parses those visible
values locally so a guessed search string can never silently produce a partial
regular-history set.

The ordinary window is the target report date. One narrow carry-over window is
also retained: a row worked on the target date whose report date is exactly one
day earlier. That is the auditable shape produced when an older regular item is
held from the afternoon aggregate and reintroduced in the following morning
aggregate.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any

from process_job import SOURCE_HEADERS, clean_text, parse_date


def canonical_header(value: Any) -> str:
    compact = "".join(clean_text(value).replace("TINY URL", "").split()).replace("*", "")
    for expected in SOURCE_HEADERS:
        if compact == "".join(expected.split()).replace("*", ""):
            return expected
    return clean_text(value)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="작업내역 표시값 범위 조회 결과를 대상일로 필터링")
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--input", help="Google Sheets bounded range 응답 JSON")
    source.add_argument("--xlsx", help="사용자가 작업내역: 으로 명시한 원본 XLSX")
    parser.add_argument("--target-date", required=True, help="YYYY-MM-DD")
    parser.add_argument("--spreadsheet-id")
    parser.add_argument("--sheet-name")
    parser.add_argument("--output", required=True)
    return parser.parse_args()


def prepare_history(raw: Any, target_date: dt.date, audit_source: dict[str, Any]) -> dict[str, Any]:
    chunks = raw.get("chunks") if isinstance(raw, dict) else None
    if not isinstance(chunks, list):
        chunks = [raw]
    scan_ranges: list[str] = []
    values: list[list[Any]] = []
    for chunk in chunks:
        chunk_values = chunk.get("values") if isinstance(chunk, dict) else None
        chunk_range = clean_text(chunk.get("range")) if isinstance(chunk, dict) else ""
        if not isinstance(chunk_values, list):
            raise ValueError("작업내역 범위 조회 JSON에 values 2차원 배열이 없습니다.")
        if not chunk_range:
            raise ValueError("작업내역 범위 조회 JSON에 실제 scan range가 없습니다.")
        scan_ranges.append(chunk_range)
        # Google Sheets may return an empty values array for a bounded trailing
        # grid chunk whose cells are all blank. Keep that queried range in the
        # source audit, but do not require a duplicate header from the empty area.
        if not chunk_values:
            continue
        if not isinstance(chunk_values[0], list):
            raise ValueError("작업내역 범위 조회 JSON에 values 2차원 배열이 없습니다.")
        if not values:
            values.extend(chunk_values)
        else:
            # Later row chunks may repeat the header. Remove it only when the
            # normalized cells actually match the first chunk's header.
            header = [canonical_header(value) for value in values[0]]
            incoming_header = [canonical_header(value) for value in chunk_values[0]]
            values.extend(chunk_values[1:] if incoming_header == header else chunk_values)
    if not values or not isinstance(values[0], list):
        raise ValueError("작업내역 범위 조회 JSON에 비어 있지 않은 헤더 행이 없습니다.")
    headers = [canonical_header(value) for value in values[0]]
    if "보도일" not in headers or "제목 (한글)" not in headers:
        raise ValueError("작업내역 범위 조회에 필수 열 보도일/제목 (한글)이 없습니다.")
    date_index = headers.index("보도일")
    work_date_index = headers.index("작업날짜") if "작업날짜" in headers else None
    previous_date = target_date - dt.timedelta(days=1)
    filtered: list[list[Any]] = []
    report_date_matches = 0
    work_date_carryovers = 0
    for row in values[1:]:
        report_date = parse_date(row[date_index]) if date_index < len(row) else None
        work_date = (
            parse_date(row[work_date_index])
            if work_date_index is not None and work_date_index < len(row)
            else None
        )
        report_date_match = report_date == target_date
        work_date_carryover = work_date == target_date and report_date == previous_date
        if not (report_date_match or work_date_carryover):
            continue
        filtered.append(row)
        report_date_matches += int(report_date_match)
        work_date_carryovers += int(work_date_carryover and not report_date_match)
    if not filtered:
        raise ValueError(f"작업내역 전체 범위에서 {target_date.isoformat()} 행을 찾지 못했습니다.")
    return {
        "source_audit": {
            "schema_version": 2,
            "retrieval_method": "bounded_range_scan",
            "value_render_option": "FORMATTED_VALUE",
            "scan_range": scan_ranges[0] if len(scan_ranges) == 1 else "",
            "scan_ranges": scan_ranges,
            "target_date": target_date.isoformat(),
            "selection_rule": "report_date_or_one_day_work_date_carryover",
            "scanned_row_count": max(0, len(values) - 1),
            "matched_row_count": len(filtered),
            "matched_by_report_date_count": report_date_matches,
            "matched_by_work_date_carryover_count": work_date_carryovers,
            **audit_source,
        },
        "values": [headers, *filtered],
    }


def xlsx_cell_value(cell: Any, cached: Any, header: str, epoch: dt.datetime) -> Any:
    """Read values and link targets only; never evaluate formulas or fetch links."""
    from openpyxl.utils.datetime import from_excel

    value = cell.value
    url_column = header in {"온라인 기사 URL", "URL (단축)"}
    if url_column and cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    if cell.data_type == "f":
        # A literal HYPERLINK needs no cached value or formula execution. Other
        # formulas must have a saved result; a label is never an article URL.
        match = re.fullmatch(
            r'=\s*(?:_xlfn\.)?HYPERLINK\(\s*"((?:[^"]|"")*)"\s*(?:[,;]\s*"(?:[^"]|"")*"\s*)?\)',
            str(value), re.IGNORECASE | re.DOTALL,
        )
        if url_column and match:
            return match.group(1).replace('""', '"')
        if url_column or cached.value is None:
            raise ValueError(f"작업내역 XLSX {cell.coordinate}: 수식을 값으로 저장한 파일이 필요합니다.")
        value = cached.value
    if cell.data_type == "e" or cached.data_type == "e":
        raise ValueError(f"작업내역 XLSX {cell.coordinate}: 오류 셀이 있습니다.")
    if header in {"보도일", "작업날짜"} and isinstance(value, (int, float)) and not isinstance(value, bool):
        value = from_excel(value, epoch)
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return value


def prepare_xlsx_history(input_path: Path, target_date: dt.date) -> dict[str, Any]:
    """Read a single unambiguous source table from a snapshot of the given XLSX."""
    from openpyxl import load_workbook

    input_path = input_path.resolve()
    if not input_path.is_file():
        raise ValueError("지정한 작업내역 XLSX 파일이 없습니다. Google Sheets로 대체하지 않습니다.")
    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("작업내역은 .xlsx 파일이어야 합니다. Google Sheets로 대체하지 않습니다.")
    if input_path.stat().st_size > 50_000_000:
        raise ValueError("작업내역 XLSX 파일이 읽기 한도를 초과했습니다.")
    snapshot = input_path.read_bytes()
    try:
        with zipfile.ZipFile(io.BytesIO(snapshot)) as archive:
            if sum(item.file_size for item in archive.infolist()) > 100_000_000:
                raise ValueError("작업내역 XLSX의 압축 해제 크기가 읽기 한도를 초과했습니다.")
        book = load_workbook(io.BytesIO(snapshot), data_only=False, keep_links=False)
        try:
            saved = load_workbook(io.BytesIO(snapshot), data_only=True, keep_links=False)
            try:
                tables = []
                for sheet in book.worksheets:
                    if sheet.max_row * 15 > 2_000_000:
                        raise ValueError("작업내역 XLSX의 행 수가 읽기 한도를 초과했습니다.")
                    for row in sheet.iter_rows(max_col=15):
                        headers = [canonical_header(cell.value) for cell in row]
                        if set(SOURCE_HEADERS).issubset(headers):
                            if any(headers.count(header) != 1 for header in SOURCE_HEADERS):
                                raise ValueError("작업내역 XLSX에 중복 헤더가 있습니다.")
                            tables.append((sheet, row[0].row, headers))
                if len(tables) != 1:
                    raise ValueError("작업내역 XLSX에서 필수 11개 헤더를 가진 표가 정확히 하나여야 합니다.")
                sheet, first_row, headers = tables[0]
                values = [headers]
                cached_sheet = saved[sheet.title]
                for row in sheet.iter_rows(min_row=first_row + 1, max_col=15):
                    values.append([
                        xlsx_cell_value(cell, cached_sheet.cell(cell.row, cell.column), headers[index], book.epoch)
                        for index, cell in enumerate(row)
                    ])
                while len(values) > 1 and all(value in (None, "") for value in values[-1]):
                    values.pop()
                sheet_name = sheet.title
            finally:
                saved.close()
        finally:
            book.close()
    except (zipfile.BadZipFile, KeyError, OSError):
        raise ValueError("작업내역 XLSX를 읽을 수 없습니다. 파일 형식과 저장 상태를 확인하세요.") from None
    scan_range = f"'{sheet_name.replace(chr(39), chr(39) * 2)}'!A{first_row}:O{first_row + len(values) - 1}"
    return prepare_history({"range": scan_range, "values": values}, target_date, {
        "schema_version": 3,
        "retrieval_method": "local_xlsx",
        "value_render_option": "XLSX_SAVED_VALUES_AND_HYPERLINKS",
        "source_file": str(input_path),
        "source_sha256": hashlib.sha256(snapshot).hexdigest(),
        "sheet_name": sheet_name,
    })


def main() -> int:
    args = parse_args()
    output_path = Path(args.output).resolve()
    input_path = Path(args.xlsx or args.input).resolve()
    if output_path == input_path:
        raise ValueError("원본 입력 파일을 출력으로 덮어쓸 수 없습니다.")
    target_date = dt.date.fromisoformat(args.target_date)
    if args.xlsx:
        if args.spreadsheet_id or args.sheet_name:
            raise ValueError("XLSX 입력에는 Google Sheets 문서·탭 인자를 함께 지정하지 않습니다.")
        payload = prepare_xlsx_history(input_path, target_date)
    else:
        if not args.spreadsheet_id or not args.sheet_name:
            raise ValueError("Google Sheets 입력에는 문서 ID와 탭 이름이 필요합니다.")
        raw = json.loads(input_path.read_text(encoding="utf-8-sig"))
        payload = prepare_history(raw, target_date, {
            "spreadsheet_id": clean_text(args.spreadsheet_id),
            "sheet_name": clean_text(args.sheet_name),
            "source_file": str(input_path),
        })
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(output_path), "matched_rows": len(payload["values"]) - 1}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
